# coding = UTF-8
import pandas as pd
import openpyxl as xl
import copy
#import treelib


def get_sites_dict():
    wb = xl.load_workbook("testdoc/Site Information.xlsx")
    ws = wb["Site Capacity(Hybrid&Packet)"]

    row_count = ws.max_row
    sites_dict = {}
    for index in range(2, row_count+1):
        site_id = ws.cell(row = index, column = 1).value
        belong_to = ws.cell(row = index, column = 11).value
        sites_dict[site_id] = belong_to
    return sites_dict


def find_path(graph, start, end, path=[]):
    path = path + [start]
    if start == end:
        return path
    if start not in graph:
        return None
    for node in graph[start]:
        if node not in path:
            newpath = find_path(graph, node, end, path)
            if newpath != None:
                return newpath
    return None


def find_path_with_connection(graph, start, end, path=[]):
    path = path + [start]
    if start == end:
        return path
    if start not in graph:
        return None
    for node in graph[start]:
        if node.site_2_id not in path:
            newpath = find_path_with_connection(graph, node.site_2_id, end, path)
            if newpath != None:
                return newpath
    return None


class LinkNode():
    def __init__(self):
        site_1_id = ''
        site_2_id = ''
        link_distance = ''
        frequency_band = ''
        protection = ''
        planned_capacity = ''
        antenna_diameter = ''
        site_1_antenna_height = ''
        site_2_antenna_height = ''
        link_availability = ''
    def swap_site_id(self):
        self.site_1_id, self.site_2_id = self.site_2_id, self.site_1_id
        self.site_1_antenna_height, self.site_2_antenna_height = \
            self.site_2_antenna_height, self.site_1_antenna_height


def generate_topology():
    wb_link = xl.load_workbook("testdoc/Link Information.xlsx")
    ws_link = wb_link['Link']
    row_count = ws_link.max_row
    link_graph = {}
    for index in range(2, row_count+1):
        link_node = LinkNode()
        link_node.site_1_id = ws_link.cell(row = index, column = 3).value
        link_node.site_2_id = ws_link.cell(row = index, column = 4).value
        link_node.link_distance = ws_link.cell(row = index, column = 11).value
        link_node.frequency_band = ws_link.cell(row = index, column = 12).value
        protection = ws_link.cell(row = index, column = 13).value
        if ws_link.cell(row = index, column = 14).value == "Yes":
            protection = protection + "XPIC"
        link_node.protection = protection
        link_node.planned_capacity = ws_link.cell(row = index, column = 15).value
        link_node.antenna_diameter = ws_link.cell(row = index, column = 16).value
        link_node.site_1_antenna_height = ws_link.cell(row = index, column = 17).value
        link_node.site_2_antenna_height = ws_link.cell(row = index, column = 18).value
        link_node.link_availability = ws_link.cell(row = index, column = 19).value
        if link_node.site_1_id not in link_graph:
            link_graph[link_node.site_1_id] = [link_node]
        else:
            link_graph[link_node.site_1_id].append(link_node)
        link_node_1 = copy.deepcopy(link_node)
        link_node_1.swap_site_id()
        if link_node_1.site_1_id not in link_graph:
            link_graph[link_node_1.site_1_id] = [link_node_1]
        else:
            link_graph[link_node_1.site_1_id].append(link_node_1)
    return link_graph

def generate_link_information():
    wb = xl.load_workbook("testdoc/City.xlsx")
    ws = wb["Technical Info"]
    row_count = ws.max_row
    sites_dict = get_sites_dict()
    link_graph = generate_topology()
    for index in range(4, row_count+1):
        site_type = ws.cell(row = index, column = 1).value
        site_id = ws.cell(row = index, column = 2).value
        if site_type == "SRAN" and site_id in sites_dict:
            fill_one_row(ws, index, sites_dict[site_id], link_graph)
    wb.save('testdoc/test.xlsx')


def fill_one_row(ws, index, belong_to, link_graph):
    site_id = ws.cell(row = index, column = 2).value
    path = find_path_with_connection(link_graph, site_id, belong_to)
    if path == None or len(path)<2:
        return
    # find the next hop to root direction
    link_node = find_node_with_name(link_graph, site_id, path[1])
    if link_node == None:
        return
    ws.cell(row = index, column = 16).value = belong_to
    ws.cell(row = index, column = 18).value = link_node.site_2_id
    ws.cell(row = index, column = 19).value =  link_node.site_1_id
    ws.cell(row = index, column = 17).value = "_".join([link_node.site_2_id, link_node.site_1_id])
    ws.cell(row = index, column = 20).value = link_node.link_distance
    ws.cell(row = index, column = 21).value = link_node.frequency_band
    ws.cell(row = index, column = 22).value = link_node.protection
    ws.cell(row = index, column = 23).value = link_node.planned_capacity
    ws.cell(row = index, column = 24).value = link_node.antenna_diameter
    ws.cell(row = index, column = 25).value = link_node.site_2_antenna_height
    ws.cell(row = index, column = 26).value = link_node.site_1_antenna_height
    ws.cell(row = index, column = 29).value = link_node.link_availability

def find_node_with_name(link_graph, site_1_id, site_2_id):
    for node in link_graph[site_1_id]:
        if node.site_2_id == site_2_id:
            return node
    return None


if __name__ == '__main__':
    """
    graph = {'A': ['B', 'C'],
             'B': ['D', 'E', 'F'],
             'C': ['H'],
             'E': ['G']}
    print('A->G', find_path(graph, 'A', 'G'))
    print('A->C', find_path(graph, 'A', 'C'))
    print('A->H', find_path(graph, 'A', 'H'))
    print('A->F', find_path(graph, 'A', 'F'))
    test_dict = {"07-06855-30-05": "07-07240-30-06",
                 "07-07240-30-06": "07-07240-30-06",
                 "07-07240-30-06": "07-07240-30-06",
                 "07-07240-30-06": "07-07240-30-06"}
    sites_dict = get_sites_dict()
    error_flag = False
    for key in test_dict:
        if key not in sites_dict or sites_dict[key] != test_dict[key]:
            error_flag = True
            print("%s belongs to %s, but the correct belonging should be %s.".format(key, sites_dict[key], test_dict[key]))
    print("test finished, result:", not error_flag)
    """
    generate_link_information()
    print('finished')
    

